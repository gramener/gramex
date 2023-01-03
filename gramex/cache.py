'''Caching utilities'''
import atexit
import contextlib
import copy
import inspect
import io
import json
import mimetypes
import os
import pandas as pd
import re
import requests
import sqlalchemy as sa

# B404:import_subprocess only developers can access this, not users
import subprocess  # nosec B404
import sys
import tempfile
import time
import tornado.ioloop
import tornado.template
from gramex.config import PathConfig
from gramex.config import app_log, merge, used_kwargs, CustomJSONDecoder, CustomJSONEncoder
from orderedattrdict import AttrDict
from queue import Queue
from threading import Thread
from tornado.concurrent import Future
from types import ModuleType
from typing import Any, List, Tuple, Union, Dict, Callable, BinaryIO
from typing_extensions import Literal
from urllib.parse import urlparse


MILLISECOND = 0.001  # in seconds
_opener_defaults = {
    'mode': 'r',
    'buffering': -1,
    'encoding': 'utf-8',
    'errors': 'strict',
    'newline': None,
    'closefd': True,
}
_markdown_defaults = {
    'output_format': 'html5',
    'extensions': [
        'markdown.extensions.codehilite',
        'markdown.extensions.extra',
        'markdown.extensions.toc',
        'markdown.extensions.meta',
        'markdown.extensions.sane_lists',
        'markdown.extensions.smarty',
    ],
}
# A set of temporary files to delete on program exit
_TEMP_FILES = set()
_ID_CACHE = set()
# In read_excel, we use range= as a parameter. Store the built-in range() to reference it
builtin_range = range


def _delete_temp_files():
    for path in _TEMP_FILES:
        if os.path.exists(path):
            os.remove(path)


atexit.register(_delete_temp_files)


def open(
    path: str,
    callback: Union[str, Callable] = None,
    transform: Callable = None,
    rel: bool = False,
    **kwargs: dict,
) -> Any:
    '''Reads a file, processes it via a callback, caches the result and returns it.

    When called again, returns the cached result unless the file has updated.

    Examples:

        >>> gramex.cache.open('data.yaml')
        >>> gramex.cache.open('data.csv')
        >>> # Load data.json as JSON into an AttrDict
        >>> gramex.cache.open('data.json', 'json', object_pairs_hook=AttrDict)
        >>> # Load data.csv as CSV into a Pandas DataFrame
        >>> gramex.cache.open('data.csv', 'csv', encoding='cp1252')

    Parameters:

        path: path to the file to open
        callback: type of file, e.g. `csv`, `json`, or callback function
        transform: function to transform the data before caching
        rel: if True, path is relative to the calling file. Default: False
        **kwargs: passed to the callback function

    `callback=` accepts these predefined types:

    - `"xlsx"`, `"xls"` or `"excel"`: reads Excel files using [gramex.cache.read_excel][].
        This supports named tables, named ranges, and ranges (e.g. `A1:B10`).
    - `"csv"`, `"table"`, `"html"`, `"sas"`, `"stata"`, `"parquet"`, `"feather"`, `"hdf"`, `"h5"`:
        reads using [Pandas](https://pandas.pydata.org/pandas-docs/stable/)
    - `"json"`: reads files using
        [`json.load()`](https://docs.python.org/3/library/json.html#json.load)
    - `"jsondata"`: reads files using
        [`pd.read_json()`](https://pandas.pydata.org/docs/reference/api/pandas.read_json.html)
    - `"yaml"`: reads files using `yaml.safe_load()`
    - `"config"`: reads files using using [gramex.config.PathConfig][].
        Same as `yaml`, but allows `import:` and variable substitution.
    - `"template"`: reads files using [`tornado.Template()`]
    - `"markdown"` or `md`: reads files using `markdown.markdown()`
    - `"text"` or `txt`: reads text files using io.open
    - `"bin"`: reads binary files using io.open

    Returns:
        The loaded file (e.g. as a DataFrame, template, etc.)

    `callback=` can be a function. For example, this returns the size of a file (cached):

        def size(path):
            return os.stat(path).st_size
        gramex.cache.open('data.fmt', size)

    To add new callbacks, use `gramex.cache.open_callback[key] = method`:

        gramex.cache.open_callback['shp'] = geopandas.read_file
        prs = gramex.cache.open('my.shp', layer='countries')

    `transform=` is an optional function that processes the data returned by
    the callback. For example:

        # Returns the count of the CSV file, updating it only when changed
        open('data.csv', 'csv', transform=lambda data: len(data))

        # After loading data.xlsx into a DataFrame, returned the grouped result
        open('data.xlsx', 'xslx', transform=lambda data: data.groupby('city')['sales'].sum())

    If `transform=` is not a function, it is used as a cache key.
    You can use this to fetch multiple cached version of the file. For example:

        >>> original = open('data.csv', 'csv', transform='v1')
        >>> same_copy = open('data.csv', 'csv', transform='v1')
        >>> new_copy = open('data.csv', 'csv', transform='v2')
        >>> original.columns = ['x', 'y', 'z']
        >>> same_copy.columns
        ['x', 'y', 'z']
        >>> new_copy.columns
        ['a', 'b', 'c']

    `rel=True` opens the path relative to the caller function's file path. If
    `D:/app/calc.py` calls `open('data.csv', 'csv', rel=True)`, the path
    is replaced with `D:/app/data.csv`.

    All ``kwargs`` are passed directly to the callback. If the callback is a predefined string
    using `io.open()`, all `io.open()` arguments are passed to `io.open()`, rest to the callback.
    '''
    # Pass _reload_status = True for testing purposes. This returns a tuple:
    # (result, reloaded) instead of just the result.
    _reload_status = kwargs.pop('_reload_status', False)
    reloaded = False
    _cache = kwargs.pop('_cache', _OPEN_CACHE)

    # Get the parent frame's filename. Compute path relative to that.
    if rel:
        path = _relpath(path)

    original_callback = callback
    if callback is None:
        callback = os.path.splitext(path)[-1][1:]
    callback_is_str = isinstance(callback, str)
    key = (
        path,
        original_callback if callback_is_str else id(callback),
        hashfn(transform),
        frozenset(((k, hashed(v)) for k, v in kwargs.items())),
    )
    cached = _cache.get(key, _FALLBACK_MEMORY_CACHE.get(key, None))
    fstat = stat(path)
    if cached is None or fstat != cached.get('stat'):
        reloaded = True
        if callable(callback):
            data = callback(path, **kwargs)
        elif callback_is_str:
            method = None
            method = open_callback.get(callback)
            if method is not None:
                data = method(path, **kwargs)
            elif original_callback is None:
                raise TypeError(f'gramex.cache.open: path "{path}" has unknown extension')
            else:
                raise TypeError(f'gramex.cache.open(callback="{callback}") is not a known type')
        else:
            raise TypeError(f'gramex.cache.open(callback=) must be a function, not {callback!r}')
        if callable(transform):
            data = transform(data)
        cached = {'data': data, 'stat': fstat}
        try:
            _cache[key] = cached
        except TypeError as e:
            # Redis / Disk caches can't pickle templates, etc. Fall back quietly to memory cache
            app_log.debug(f'gramex.cache.open: {e} on {callback}. Using fallback memory cache')
            _FALLBACK_MEMORY_CACHE[key] = cached
        except ValueError:
            size = sys.getsizeof(data)
            app_log.exception(
                f'gramex.cache.open: {type(_cache):s} cannot cache {size} bytes. '
                + 'Increase cache.memory.size in gramex.yaml'
            )
        except Exception:
            app_log.exception(f'gramex.cache.open: {type(_cache)} cannot cache {data!r}')

    result = cached['data']
    return (result, reloaded) if _reload_status else result


def read_excel(
    io: Union[str, BinaryIO],
    sheet_name: Union[str, int] = 0,
    table: str = None,
    name: str = None,
    range: str = None,
    header: Union[None, int, List[int]] = ...,
    **kwargs: dict,
) -> pd.DataFrame:
    '''Read data from an XLSX as a DataFrame using `openpyxl`.

    Examples:

        >>> gramex.cache.read_excel('data.xlsx', sheet_name='Sheet1')
        >>> gramex.cache.read_excel('data.xlsx', sheet_name=0)
        >>> gramex.cache.read_excel('data.xlsx', table='Table1')
        >>> gramex.cache.read_excel('data.xlsx', name='NamedRange')
        >>> gramex.cache.read_excel('data.xlsx', range='A1:D10', header=[0, 1])

    Parameters:

        io: path or file-like object pointing to an Excel file
        sheet_name: sheet to load data from. Sheet names are specified as strings.
            Integers pick zero-indexed sheet position. default: 0
        table: Worksheet table to load from sheet, e.g. `'Table1'`
        name: Defined name to load from sheet, e.g. `'MyNamedRange'`
        range: Cell range to load from sheet, e.g. `'A1:C10'`
        header: Row (0-indexed) to use for the column labels. A list of integers is combined into
            a MultiIndex. Use None if there is no header.
        **kwargs: If neither `table`, nor `name`, nor `range` is specified, loads entire
            sheet via `pd.read_excel`, passing the remaining kwargs.

    Returns:
        The loaded DataFrame

    Note: `table` takes priority over `name` takes priority over `range`.
    '''
    if not any((range, name, table)):
        # Pandas defaults to xlrd, but we prefer openpyxl
        kwargs.setdefault('engine', 'openpyxl')
        return pd.read_excel(
            io, sheet_name=sheet_name, header=0 if header is ... else header, **kwargs
        )

    import openpyxl

    wb = openpyxl.load_workbook(io, data_only=True)
    # Pick a SINGLE sheet using sheet_name -- it can be an int or a str
    ws = wb[wb.sheetnames[sheet_name] if isinstance(sheet_name, int) else sheet_name]
    # Get the data range to be picked
    if table is not None:
        range = ws.tables[table].ref
        # Tables themselves specify whether they have a column header. Use this as default
        if header is ...:
            header = list(builtin_range(ws.tables[table].headerRowCount))
    elif name is not None:
        # If the name is workbook-scoped, get it directly
        defined_name = wb.defined_names.get(name)
        # Else, if it's sheet-scoped, get it related to the sheet
        if defined_name is None:
            defined_name = wb.defined_names.get(name, wb.sheetnames.index(ws.title))
        # Raise an error if we can't find it
        if defined_name is None:
            raise ValueError(f'{io}: missing name {name} in sheet {sheet_name}')
        # Note: This only works if it's a cell range. If we create a named range inside a table,
        # Excel may store this as =Table[[#All],[Col1]:[Col5]], which isn't a valid range.
        # Currently, we ignore that, and assumed that the name is like Sheet1!A1:C10
        range = defined_name.attr_text.split('!')[-1]

    vals = ws[range]
    if isinstance(vals, tuple):
        data = pd.DataFrame([[cell.value for cell in row] for row in vals])
    else:
        data = pd.DataFrame([[vals.value]])
    # Header defaults to 0 if undefined. If it's not None, apply the header
    header = 0 if header is ... else header
    if header is not None:
        data = (
            data.T.set_index(header)
            .T.reset_index(  # Set header rows as column names
                drop=True
            )  # Drop index with "holes" where headers were
            .rename_axis(  # Column name has header index (e.g. 0). Drop it
                [None] * len(header) if isinstance(header, (list, tuple)) else None, axis=1
            )
        )
    return data.infer_objects()  # Convert data types


def save(
    data: pd.DataFrame,
    url: str,
    rel: bool = False,
    callback: Union[str, Callable] = None,
    **kwargs: dict,
) -> None:
    '''Saves a Pandas DataFrame into file at url.

    Examples:

        >>> gramex.cache.save(pd.DataFrame({'x': [1, 2]}), 'sample.csv')
        >>> gramex.cache.save(pd.DataFrame({'x': [1, 2]}), 'sample.xlsx', sheet_name='data')

    Parameters:
        data: Pandas dataframe which has to be saved to a file.
        url: path where the dataframe has to be saved.
        rel: if True, path is relative to the calling file. Default: False
        callback: Almost the same as for [gramex.cache.open][]. It can
            be `json`, `csv`, `xlsx`, `hdf`, `html`, `stata` or
            a function that accepts the filename and any other arguments.
        **kwargs: Other keyword arguments are passed directly to the callback.
    '''
    if callback is None:
        callback = os.path.splitext(url)[-1][1:]
    if callable(callback):
        return callback(data, url, **kwargs)
    elif callback in _SAVE_CALLBACKS:
        url = _relpath(url) if rel else url
        method = getattr(data, _SAVE_CALLBACKS[callback])
        return method(url, **(used_kwargs(method, kwargs)[0]))
    else:
        raise TypeError(f'gramex.cache.save(callback="{callback}") is unknown')


def query(
    sql: str,
    engine: sa.engine.base.Engine,
    state: Union[str, Callable, List[str], None] = None,
    **kwargs: dict,
):
    '''Read SQL query or database table into a DataFrame, cached.

    It uses `state` to determine whether to run the query again or not.

    Examples:

        >>> engine = sqlalchemy.create_engine('sqlite:///path/to/file.db')
        >>> gramex.cache.query('SELECT * FROM table', engine, state='SELECT max(date) FROM table')

    Parameters:

        sql: SQL query or table name to read from the database.
        engine: SQLAlchemy engine to read from.
        state: An SQL query, function or list of table names to check if the data has changed.
        kwargs: Other keyword arguments are passed directly to `pandas.read_sql`.

    `state` can be a:

    1. `string`, e.g. `SELECT max(date) FROM updated`. This must be as a lightweight SQL query.
        If the result changes, the SQL query is re-run.
    2. `function`, e.g. `lambda v: os.stat('transfer.log').st_mtime`.
        If the result changes, the SQL query is re-run.
    3. list of table names, e.g. `["db.table1", "db.table2"]`.
        If any of these tables have changed, the SQL query is re-run. **EXPERIMENTAL**
    4. `None`: the default. The query is always re-run and not cached.
    '''
    # Pass _reload_status = True for testing purposes. This returns a tuple:
    # (result, reloaded) instead of just the result.
    _reload_status = kwargs.pop('_reload_status', False)
    reloaded = False
    _cache = kwargs.pop('_cache', _QUERY_CACHE)
    store_cache = True

    if isinstance(state, (list, tuple)):
        try:
            status = _table_status(engine, tuple(state))
        except KeyError as e:
            # Unknown SQLAlchemy dialects raise a KeyError.
            # Warn and don't cache the table. (Otherwise, no new dialects can be used.)
            app_log.warning(e.args[0] if len(e.args) > 0 else 'gramex.cache.query: state failed')
            status, store_cache = object(), False
    elif isinstance(state, str):
        status = pd.read_sql(state, engine).to_dict(orient='list')
    elif callable(state):
        status = state()
    elif state is None:
        # Create a new status every time, so that the query is always re-run
        status, store_cache = object(), False
    else:
        raise TypeError(f'gramex.cache.query(state=) must be a table list/query/fn, not {state!r}')

    key = (str(sql), cache_key(kwargs.get('params', {})), engine.url)
    if key in _cache and _cache[key]['status'] == status:
        result = _cache[key]['data']
    else:
        app_log.debug(
            f'gramex.cache.query: {sql}. engine: {engine}. state: {state}. kwargs: {kwargs}'
        )
        result = pd.read_sql(sql, engine, **kwargs)
        if store_cache:
            _cache[key] = {
                'data': result,
                'status': status,
            }
        reloaded = True

    return (result, reloaded) if _reload_status else result


def stat(path: str) -> Union[Tuple[float, int], Tuple[None, None]]:
    '''Returns a file's modified time and size. Used to check if a file has changed.

    Examples:

        >>> gramex.cache.stat('gramex.yaml')
        (1654149106.1422858, 7675)
        >>> gramex.cache.stat('non-existent-file')
        (None, None)

    Parameters:
        path: Absolute file path/Path relative to gramex root folder

    Returns:
        The last modified time and file size, or `(None, None)` if the file does not exist.
    '''
    if os.path.exists(path):
        stat = os.stat(path)
        return (stat.st_mtime, stat.st_size)
    return (None, None)


def _json_dump(obj: Any) -> str:
    '''Dumps an object to JSON, sorted by key for stability.'''
    return json.dumps(
        obj, sort_keys=True, separators=(',', ':'), ensure_ascii=True, cls=CustomJSONEncoder
    )


def cache_key(*args: List[Any]) -> str:
    '''Converts arguments into a stable string suitable for use as a cache key.

    Examples:

        >>> t1 = gramex.cache.cache_key(tweet1)
        >>> data[t1] = tweet1
        >>> t2 = gramex.cache.cache_key(tweet2)
        >>> data[t2] = tweet2

    Parameters:

        args: Any number of arguments to serialized with JSON.

    Returns:
        A JSON dump of the arguments, sorted by key for stability.
    '''
    return _json_dump(args)


def hashed(val: Any) -> str:
    '''Return the hashed value of val. If not possible, return None'''
    try:
        hash(val)
        return val
    except TypeError:
        try:
            return cache_key(val)
        except Exception:
            return None


# reload_module() stores its cache here
_MODULE_CACHE = {}


def reload_module(*modules: List[ModuleType]) -> None:
    '''Reloads one or more modules if they have changed on disk.

    Examples:

        >>> import mymodule1
        >>> import mymodule2
        >>> reload_module(mymodule1, mymodule2)

    Parameters:

        *modules: Pass the module which has to reload.

    This is useful when developing templates that import modules. Use:

        {% import mymodule %}
        {% import gramex.cache %}
        {% gramex.cache.reload_module(mymodule) %}
        {# ... use mymodule ... #}

    If `mymodule.py` changes, the template will automatically reload it.
    '''
    for module in modules:
        name = getattr(module, '__name__', None)
        path = getattr(module, '__file__', None)
        # sys.__file__ does not exist, but don't raise a warning. You can't reload it
        if name in {'sys'}:
            continue
        if name is None or path is None or not os.path.exists(path):
            app_log.warning(f'Cannot locate path for module "{name}". Got path: {path}')
            continue
        # The first time, don't reload it. Thereafter, if it's older or resized, reload it
        fstat = stat(path)
        if fstat != _MODULE_CACHE.get(name, fstat):
            app_log.info(f'Reloading module {name}')
            import importlib

            importlib.reload(module)
        _MODULE_CACHE[name] = fstat


def urlfetch(url: str, info: bool = False, **kwargs: dict) -> Union[str, Dict]:
    '''Fetch the content in the url and return a file path where it is downloaded.

    Examples:

        >>> gramex.cache.urlfetch('https://gramener.com/gramex/guide/mlhandler/titanic')
        '/path/to/tmpfile.json'

    Parameters:
        url: A http, https or file path
        info: True if metadata of the requested file is required. If true, it
            returns a dict with (filename), r (request) url, ext (extension), content_type.
        **kwargs: Any other keyword arguments are passed to requests.get.

    Returns:
        Filepath where the file is downloaded.

    Note:

    - When Gramex exits, the files are automatically deleted.
    - The extension is based on the URL's Content-Type HTTP header.
    - If `url` is the path to a file, it returns the `url` itself.
    '''
    urlparts = urlparse(url)
    if urlparts.scheme not in {'http', 'https'}:  # url is a filepath
        if info:
            ext = os.path.splitext(url)[1]
            content_type = mimetypes.guess_type(url, strict=True)[0]
            return {'name': url, 'r': None, 'url': None, 'ext': ext, 'content_type': content_type}
        else:
            return url
    r = requests.get(url, **kwargs)
    if 'Content-Type' in r.headers:
        content_type = r.headers['Content-Type'].split(';')[0]
        ext = mimetypes.guess_extension(content_type, strict=False)
    else:
        ext = os.path.splitext(urlparts.path)[1]
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as handle:
        for chunk in r.iter_content(chunk_size=16384):
            handle.write(chunk)
    _TEMP_FILES.add(handle.name)
    if info:
        return {'name': handle.name, 'r': r, 'url': url, 'ext': ext, 'content_type': content_type}
    else:
        return handle.name


class Subprocess:
    '''Wraps `subprocess.Popen` to run in a Tornado IOLoop.

    [`tornado.process.Subprocess()`](https://www.tornadoweb.org/en/stable/process.html)
    [fails on Windows](https://github.com/tornadoweb/tornado/issues/1585).
    This is a threaded alternative.

    Examples:

        Run a program async and wait for it to execute. Then get its output:

        >>> stdout, stderr = yield Subprocess(['ls', '-la']).wait_for_exit()

        Run a program async and send each line to the handler as it writes:

        >>> yield Subprocess(
        ...     ['ls', '-la'],                  # Run 'ls -la'
        ...     buffer_size='line',             # Buffer output line by line
        ...     stream_stdout=handler.write,    # Send output to handler.write(line)
        ...     stream_stderr=handler.write,    # Send errors to handler.write(line)
        ... )

        Run a program async and appends output into a list:

        >>> proc = Subprocess(
        ...     ['ls', '-la'],
        ...     buffer_size='line',
        ...     stream_stdout='list_out',       # Append output to self.list_out
        ...     stream_stderr='list_err',       # Append errors to self.list_err
        ... )
        ... output = proc.list_out[-10:]        # Return last 10 lines of output
        ... yield proc.wait_for_exit()          # Wait until application is done

        Run a program async and appends output into a queue:

        >>> proc = Subprocess(
        ...     ['ls', '-la'],                  # Run 'ls -la'
        ...     buffer_size='line',             # Buffer output line by line
        ...     stream_stdout='queue_out',      # Save output in proc.out queue
        ...     stream_stderr='queue_err',      # Save errors in proc.err queue
        ... )
        ... output = proc.queue_out.get_nowait()    # Returns first line of output
        ... yield proc.wait_for_exit()              # Wait until application is done

        To write to multiple streams, pass a list:

        >>> proc = Subprocess(
        ...     args,
        ...     buffer_size='line',
        ...     stream_stdout=[handler.write, 'list_out', 'queue_out', my_callback],
        ...     stream_stderr=[handler.write, 'list_err', 'queue_err', my_callback],
        ...     **kwargs
        ... )
        ... yield proc.wait_for_exit()

        To check the process return code, use `.proc` which has the `Popen`
        object:

        >>> if proc.proc.returncode:
        ...     raise Exception('Process failed with return code %d', proc.proc.returncode)

    [Discussion](http://stackoverflow.com/a/4896288/100904)
    '''

    def __init__(
        self,
        args: List[str],
        stream_stdout: List[Union[Callable, str]] = (),
        stream_stderr: List[Union[Callable, str]] = (),
        buffer_size: Union[str, int] = 0,
        **kwargs: dict,
    ):
        '''
        Parameters:
            args: command line arguments passed as a list to Subprocess
            stream_stdout: optional list of write methods - called when stdout has data
            stream_stderr: optional list of write methods - called when stderr has data
            buffer_size: 'line' to write line by line, any int for chunk size
            **kwargs: additional kwargs passed to subprocess.Popen

        `stream_stdout` and `stream_stderr` can be:

        - a function that accept a byte string. Called as stdout/stderr are buffered
        - OR a string starting with `list_` or `queue_`. Appends buffered output
        - OR a list of any of the above
        - OR an empty list. In this case, `.wait_for_exit()` returns a tuple with
        `stdout` and `stderr` as a tuple of byte strings.
        '''
        self.args = args

        # self.proc.stdout & self.proc.stderr are streams with process output
        kwargs['stdout'] = kwargs['stderr'] = subprocess.PIPE

        # On UNIX, close all file descriptors except 0, 1, 2 before child
        # process is executed. I've no idea why. Copied from
        # http://stackoverflow.com/a/4896288/100904
        kwargs['close_fds'] = 'posix' in sys.builtin_module_names

        # B603:subprocess_without_shell_equals_true: only developers can access this, not users
        self.proc = subprocess.Popen(args, **kwargs)  # nosec B603
        self.thread = {}  # Has the running threads
        self.future = {}  # Stores the futures indicating stream close
        self.loop = _get_current_ioloop()

        # Buffering has 2 modes. buffer_size='line' reads and writes line by line
        # buffer_size=<number> reads in byte chunks. Define the appropriate method
        if hasattr(buffer_size, 'lower') and 'line' in buffer_size.lower():

            def _write(stream, callbacks, future, retval):
                '''Call callbacks with content from stream. On EOF mark future as done'''
                while True:
                    content = stream.readline()
                    if len(content) > 0:
                        if isinstance(content, str):
                            content = content.encode('utf-8')
                        for callback in callbacks:
                            callback(content)
                    else:
                        stream.close()
                        break
                while self.proc.poll() is None:
                    time.sleep(MILLISECOND)
                self.loop.add_callback(future.set_result, retval())

        else:
            # If the buffer size is 0 or negative, use the default buffer size to read
            if buffer_size <= 0:
                buffer_size = io.DEFAULT_BUFFER_SIZE

            def _write(stream, callbacks, future, retval):
                '''Call callbacks with content from stream. On EOF mark future as done'''
                while True:
                    content = stream.read(buffer_size)
                    size = len(content)
                    if size > 0:
                        if isinstance(content, str):
                            content = content.encode('utf-8')
                        for callback in callbacks:
                            # This may raise a ValueError: write to closed file.
                            # TODO: decide how to handle it.
                            callback(content)
                    if size < buffer_size:
                        stream.close()
                        break
                while self.proc.poll() is None:
                    time.sleep(MILLISECOND)
                self.loop.add_callback(future.set_result, retval())

        callbacks_lookup = {'stdout': stream_stdout, 'stderr': stream_stderr}
        for stream in ('stdout', 'stderr'):
            callbacks = callbacks_lookup[stream]
            # If stream_stdout or stream_stderr are not defined, construct a
            # BytesIO and return its value when the stream is closed
            if not callbacks:
                ret_stream = io.BytesIO()
                callbacks = [ret_stream.write]
                retval = ret_stream.getvalue
            else:
                retval = lambda: b''  # noqa
            # If stream_stdout or stream_stderr has 'out' or 'err', create these
            # as queue attributes (self.out, self.err)
            callbacks = list(callbacks) if isinstance(callbacks, list) else [callbacks]
            for index, method in enumerate(callbacks):
                if isinstance(method, str):
                    if method.startswith('list_'):
                        if hasattr(self, method):
                            callbacks[index] = getattr(self, method).append
                        else:
                            log = []
                            setattr(self, method, log)
                            callbacks[index] = log.append
                    elif method.startswith('queue_'):
                        if hasattr(self, method):
                            callbacks[index] = getattr(self, method).put
                        else:
                            log = Queue()
                            setattr(self, method, log)
                            callbacks[index] = log.put
                    else:
                        raise ValueError(f'Invalid stream_{stream}: {method}')
            self.future[stream] = future = Future()
            # Thread writes from self.proc.stdout / stderr to appropriate callbacks
            self.thread[stream] = t = Thread(
                target=_write,
                name=f'cache.Subprocess: {args}',
                args=(getattr(self.proc, stream), callbacks, future, retval),
            )
            t.daemon = True  # Thread dies with the program
            t.start()

    def wait_for_exit(self):
        '''
        Returns futures for (stdout, stderr). To wait for the process to complete, use:

            stdout, stderr = yield proc.wait_for_exit()
        '''
        return [self.future['stdout'], self.future['stderr']]


_daemons = {}


def daemon(
    args: Union[str, List[str]],
    restart: int = 1,
    first_line: Union[str, re.Pattern, Callable, None] = None,
    timeout: float = 5,
    stream: Union[bool, Callable] = True,
    buffer_size: Union[int, Literal['line']] = 'line',
    **kwargs,
) -> Future:
    '''Run a long-running process in the background, and wait till it starts properly.

    This is used to run & cache servers like Node web servers, Elasticsearch, etc. in Tornado.

    Examples:

        ```python
        @tornado.gen.coroutine
        def fetch_from_server(handler):
            server = yield gramex.cache.daemon(
                ['python', '-m', 'http_server', '8000']
                restart=3,
                first_line=re.compile(r'Serving HTTP on .* port 8000'),
                timeout=5,
                stream=sys.stdout.write,
                buffer_size='line',
            )
            return requests.get('http://localhost:8000/').text
        ```

        The first time `fetch_from_server` is called, it starts a Python web server on port 8000.
        Subsequent calls use the same server.

    Parameters:

        args: command to run. Can be a list of strings or a string. See
            [subprocess.Popen](https://docs.python.org/3/library/subprocess.html#subprocess.Popen)
        restart: number of times to restart the process if it fails. Set to 0 to never restart.
        first_line: wait until the first line of output matches this string, regex or function
        timeout: number of seconds to wait first line to appear
        stream: if True, write process output to stderr. If a function, call it with each line
            (e.g. `sys.stdout.write`)
        buffer_size: if 'line', flush stream on every new line. If an integer, flush stream
            after every `buffer_size` bytes.

    Returns:
        A future that resolves to a [gramex.cache.Subprocess][] object.

    1. If we have already called [Subprocess] with the same arguments,
        re-use the same instance.
    2. Send the process STDOUT and STDERR to this application's STDERR. This
        makes it easy to see what errors the application reports.
    3. Supports retry attempts.
    4. Checks if the first line of output is a matches a string / re -- ensuring
        that the application started properly.
    '''
    arg_str = args if isinstance(args, str) else ' '.join(args)
    # TODO: stream_stderr and stream_stdout are NOT likely to be serializable! Exclude those
    try:
        key = cache_key(arg_str, kwargs)
    except (TypeError, ValueError):
        app_log.error('daemon args must be JSON serializable')
        raise
    # Send the stdout and stderr to (a) stderr AND to (b) a local queue we read
    queue = Queue(maxsize=10)
    for channel in ('stream_stdout', 'stream_stderr'):
        if channel not in kwargs:
            kwargs[channel] = []
        elif not isinstance(kwargs[channel], list):
            kwargs[channel] = [kwargs[channel]]
        if first_line:
            kwargs[channel].append(queue.put)
        if stream is True:
            kwargs[channel].append(sys.stderr.buffer.write)
        elif callable(stream):
            kwargs[channel].append(stream)
    # Buffer by line by default. This is required for the first_line check, not otherwise.
    kwargs['buffer_size'] = buffer_size
    # started is set if we actually call Subprocess as part of this function
    started = False

    # If process was never started, start it
    if key not in _daemons:
        # B404:import_subprocess only developers can access this, not users
        started = _daemons[key] = Subprocess(args, **kwargs)  # nosec B404

    # Ensure that process is running. Restart if required
    proc = _daemons[key]
    restart = int(restart)
    while proc.proc.returncode is not None and restart > 0:
        restart -= 1
        # B404:import_subprocess only developers can access this, not users
        proc = started = _daemons[key] = Subprocess(args, **kwargs)  # nosec B404
    if proc.proc.returncode is not None:
        raise RuntimeError(f'Error {proc.proc.returncode} starting {arg_str}')
    if started:
        app_log.info(f'Started: {arg_str}')

    future = Future()
    # If process was started, wait until it has initialized. Else just return the proc
    if first_line and started:
        if isinstance(first_line, str):

            def check(proc):
                actual = queue.get(timeout=timeout).decode('utf-8')
                if first_line not in actual:
                    raise AssertionError(f'{arg_str}: first line is "{actual}" not "{first_line}"')

        elif isinstance(first_line, re.Pattern):

            def check(proc):
                actual = queue.get(timeout=timeout).decode('utf-8')
                if not first_line.search(actual):
                    raise AssertionError(f'{arg_str}: wrong first line: {actual}')

        elif callable(first_line):
            check = first_line
        loop = _get_current_ioloop()

        def checker(proc):
            try:
                check(proc)
            except Exception as e:
                loop.add_callback(future.set_exception, e)
            else:
                loop.add_callback(future.set_result, proc)

        proc._check_thread = t = Thread(target=checker, args=(proc,))
        t.daemon = True  # Thread dies with the program
        t.start()
    else:
        future.set_result(proc)
    return future


def get_store(type, **kwargs):
    if type == 'memory':
        return KeyStore(**kwargs)
    elif type == 'sqlite':
        return SQLiteStore(**kwargs)
    elif type == 'json':
        return JSONStore(**kwargs)
    elif type == 'redis':
        return RedisStore(**kwargs)
    elif type == 'hdf5':
        return HDF5Store(**kwargs)
    else:
        raise NotImplementedError(f'Store type: {type} not implemented')


class KeyStore:
    '''
    Base class for persistent dictionaries. (But KeyStore is not persistent.)

        >>> store = KeyStore()
        >>> value = store.load(key, None)   # Load a value. It's like dict.get()
        >>> store.dump(key, value)          # Save a value. It's like dict.set(), but doesn't flush
        >>> store.flush()                   # Saves to disk
        >>> store.close()                   # Close the store

    You can initialize a KeyStore with a `flush=` parameter. The store is
    flushed to disk via `store.flush()` every `flush` seconds.

    If a `purge=` is provided, the data is purged of missing values every
    `purge` seconds. You can provide a custom `purge_keys=` function that
    returns an iterator of keys to delete if any.

    When the program exits, `.close()` is automatically called.
    '''

    def __init__(self, flush=None, purge=None, purge_keys=None, **kwargs):
        self.store = {}
        if callable(purge_keys):
            self.purge_keys = purge_keys
        elif purge_keys is not None:
            app_log.error('KeyStore: purge_keys=%r invalid. Must be function(dict)', purge_keys)
        # Periodically flush and purge buffers
        if flush is not None:
            tornado.ioloop.PeriodicCallback(self.flush, callback_time=flush * 1000).start()
        if purge is not None:
            tornado.ioloop.PeriodicCallback(self.purge, callback_time=purge * 1000).start()
        # Call close() when Python gracefully exits
        atexit.register(self.close)

    def keys(self):
        '''Return all keys in the store'''
        return self.store.keys()

    def load(self, key, default=None):
        '''Same as `store.get(key)`, but it's called `load()` to indicate persistence'''
        key = self._escape(key)
        return self.store.get(key, {} if default is None else default)

    def dump(self, key, value):
        '''Same as `store[key] = value`'''
        key = self._escape(key)
        self.store[key] = value

    def _escape(self, key):
        # Converts key into a unicode string (interpreting byte-string keys as UTF-8)
        return str(key, encoding='utf-8') if isinstance(key, bytes) else str(key)

    @staticmethod
    def purge_keys(data):
        return [key for key, val in data.items() if val is None]

    def flush(self):
        '''Write to disk'''
        pass

    def purge(self):
        '''Delete empty keys and flush'''
        for key in self.purge_keys(self.store):
            # If the key was already removed from store, ignore
            with contextlib.suppress(KeyError):
                del self.store[key]
        self.flush()

    def close(self):
        '''Flush and close all open handles'''
        raise NotImplementedError()


class RedisStore(KeyStore):
    '''
    A KeyStore that stores data in a Redis database. Typical usage:

        >>> store = RedisStore('localhost:6379:1:password=x:...')     # host:port:db:params
        >>> value = store.load(key)
        >>> store.dump(key, value)

    The path in the constructor contains parameters separated by colon (:):

    - `host`: the Redis server location (default: localhost)
    - `port`: the Redis server port (default: 6379)
    - `db`: the Redis server DB number (default: 0)
    - zero or more parameters passed to StrictRedis (e.g. password=abc)

    Values are encoded as JSON using gramex.config.CustomJSONEncoder (thus
    handling datetime.) Keys are JSON encoded.
    '''

    def __init__(self, path=None, *args, **kwargs):
        super(RedisStore, self).__init__(*args, **kwargs)
        from gramex.services.rediscache import get_redis

        self.store = get_redis(path, decode_responses=True, encoding='utf-8')

    def load(self, key, default=None):
        result = self.store.get(key)
        if result is None:
            return default
        try:
            return json.loads(result, object_pairs_hook=AttrDict, cls=CustomJSONDecoder)
        except ValueError:
            app_log.error(f'RedisStore("{self.store}").load("{key}") is not JSON "{result!r}"')
            return default

    def dump(self, key, value):
        if value is None:
            self.store.delete(key)
        else:
            value = _json_dump(value)
            self.store.set(key, value)

    def close(self):
        pass

    def purge(self):
        app_log.debug(f'Purging {self.store}')
        # TODO: optimize item retrieval
        items = {key: self.load(key, None) for key in self.store.keys()}
        for key in self.purge_keys(items):
            self.store.delete(key)


class SQLiteStore(KeyStore):
    '''
    A KeyStore that stores data in a SQLite file. Typical usage:

        >>> store = SQLiteStore('file.db', table='store')
        >>> value = store.load(key)
        >>> store.dump(key, value)

    Values are encoded as JSON using gramex.config.CustomJSONEncoder (thus
    handling datetime.) Keys are JSON encoded.
    '''

    def __init__(self, path, table='store', *args, **kwargs):
        super(SQLiteStore, self).__init__(*args, **kwargs)
        self.path = _create_path(path)
        from sqlitedict import SqliteDict

        self.store = SqliteDict(
            self.path,
            tablename=table,
            autocommit=True,
            encode=_json_dump,
            decode=lambda v: json.loads(v, object_pairs_hook=AttrDict, cls=CustomJSONDecoder),
        )

    def close(self):
        self.store.close()

    def flush(self):
        super(SQLiteStore, self).flush()
        self.store.commit()

    def keys(self):
        # Keys need to be escaped
        return (self._escape(key) for key in self.store.keys())

    def purge(self):
        app_log.debug(f'Purging {self.path}')
        super(SQLiteStore, self).purge()


class HDF5Store(KeyStore):
    '''
    A KeyStore that stores data in a HDF5 file. Typical usage:

        >>> store = HDF5Store('file.h5', flush=15)
        >>> value = store.load(key)
        >>> store.dump(key, value)

    Internally, it uses HDF5 groups to store data. Values are encoded as JSON
    using gramex.config.CustomJSONEncoder (thus handling datetime.) Keys are JSON
    encoded, and '/' is escaped as well (since HDF5 groups treat / as subgroups.)
    '''

    def __init__(self, path, *args, **kwargs):
        super(HDF5Store, self).__init__(*args, **kwargs)
        self.path = _create_path(path)
        self.changed = False
        import h5py

        # h5py.File fails with OSError: Unable to create file (unable to open file: name =
        # '.meta.h5', errno = 17, error message = 'File exists', flags = 15, o_flags = 502)
        # TODO: identify why this happens and resolve it.
        self.store = h5py.File(self.path, 'a')

    def load(self, key, default=None):
        # Keys cannot contain / in HDF5 store. Escape it
        key = self._escape(key).replace('/', '\t')
        result = self.store.get(key, None)
        if result is None:
            return default
        result = result[()]
        try:
            return json.loads(result, object_pairs_hook=AttrDict, cls=CustomJSONDecoder)
        except ValueError:
            app_log.error(f'HDF5Store("{self.path}").load("{key}") is not JSON ("{result!r}")')
            return default

    def dump(self, key, value):
        key = self._escape(key)
        # TODO: BUG. Rewrite like JSONStore.dump()
        if self.store.get(key) != value:
            if key in self.store:
                del self.store[key]
            self.store[key] = _json_dump(value)
            self.changed = True

    def _escape(self, key):
        # Converts key into a unicode string (interpreting byte-string keys as UTF-8).
        # HDF5 does not accept / in key names. Replace those with tabs.
        key = str(key, encoding='utf-8') if isinstance(key, bytes) else str(key)
        return key.replace('/', '\t')

    def keys(self):
        # Keys cannot contain / in HDF5 store. Unescape it
        return (key.replace('\t', '/') for key in self.store.keys())

    def flush(self):
        super(HDF5Store, self).flush()
        if self.changed:
            app_log.debug(f'Flushing {self.path}')
            self.store.flush()
            self.changed = False

    def purge(self):
        # Load all keys into self.store. Delete what's required. Save.
        self.flush()
        changed = False
        items = {
            key: json.loads(val[()], object_pairs_hook=AttrDict, cls=CustomJSONDecoder)
            for key, val in self.store.items()
        }
        for key in self.purge_keys(items):
            del self.store[key]
            changed = True
        if changed:
            app_log.debug(f'Purging {self.path}')
            self.store.flush()

    def close(self):
        try:
            self.store.close()
        # h5py.h5f.get_obj_ids often raises a ValueError: Not a file id.
        #   This is presumably if the file handle has been closed. Log & ignore.
        # import h5py may fil. If so, self.store is a dict, and has no .close().
        #   This raises an AttributeError. Log & ignore
        except (AttributeError, ValueError) as e:
            app_log.debug(f'HDF5Store("{self.path}").close() error ({e}) ignored')


class JSONStore(KeyStore):
    '''
    A KeyStore that stores data in a JSON file. Typical usage:

        >>> store = JSONStore('file.json', flush=15)
        >>> value = store.load(key)
        >>> store.dump(key, value)

    This is less efficient than HDF5Store for large data, but is human-readable.
    They also cannot support multiple instances. Only one JSONStore instance
    is permitted per file.
    '''

    def __init__(self, path, *args, **kwargs):
        super(JSONStore, self).__init__(*args, **kwargs)
        self.path = _create_path(path)
        self._init_store(self._read_json())

    def _init_store(self, contents):
        self.store = contents
        self._original = copy.deepcopy(self.store)  # copy of original contents
        self.changed = False  # boolean: has the store contents changed?
        self.update = {}  # all key-values added since flush

    def _read_json(self):
        try:
            with io.open(self.path) as handle:
                return json.load(handle, cls=CustomJSONDecoder)
        except (IOError, ValueError):
            return {}

    def _write_json(self, data):
        json_value = _json_dump(data)
        with io.open(self.path, 'w') as handle:
            handle.write(json_value)

    def dump(self, key, value):
        '''Same as store[key] = value'''
        key = self._escape(key)
        self.store[key] = value
        # Update contents only if the value is different from the original
        if self._original.get(key) != value:
            self.update[key] = value
            self.changed = True

    def flush(self, purge=False):
        super(JSONStore, self).flush()
        if getattr(self, 'changed', False) or purge:
            app_log.debug(f"{'Purging' if purge else 'Flushing'} {self.path}")
            # Don't dump contents. That can overwrite other instances' updates.
            # Instead: read, apply updates, and save.
            store = self._read_json()
            store.update(self.update)
            for key in self.purge_keys(store):
                del store[key]
            self._write_json(store)
            self._init_store(store)

    def purge(self):
        self.flush(purge=True)

    def close(self):
        try:
            self.flush()
        # This has happened when the directory was deleted. Log & ignore.
        except OSError:
            app_log.exception(f'Cannot flush {self.path}')


def hashfn(fn):
    # Returns a unique hash value for the function.
    # id() returns a unique value for the lifetime of an object.
    # To ensure that ID is not re-cycled, cache object, so it's never released.
    _ID_CACHE.add(fn)
    return id(fn)


def opener(callback, read=False, **open_kwargs):
    '''
    Converts any function that accepts a string or handle as its parameter into
    a function that takes the first parameter from a file path.

    Examples:

        >>> jsonload = opener(json.load)
        >>> jsonload('x.json')      # opens x.json and runs json.load(handle)
        >>> gramex.cache.open('x.json', jsonload)   # Loads x.json, cached

        >>> # read=True parameter passes the contents (not handle) to the function
        >>> template = opener(string.Template, read=True)
        >>> template('abc.txt').substitute(x=val)
        >>> gramex.cache.open('abc.txt', template).substitute(x=val)

        >>> # If read=True, callback may be None. The result of .read() is passed as-is
        >>> text = opener(None, read=True)
        >>> gramex.cache.open('abc.txt', text)

    Keyword arguments applicable for `io.open` are passed to `io.open`. These
    default to `io.open(mode='r', buffering=-1, encoding='utf-8',
    errors='strict', newline=None, closefd=True)`. All other arguments and
    keyword arguments are passed to the callback (e.g. to `json.load`).

    When reading binary files, pass `mode='rb', encoding=None, errors=None`.
    '''
    merge(open_kwargs, _opener_defaults, 'setdefault')
    if read:
        # Pass contents to callback
        def method(path, **kwargs):
            open_args = {key: kwargs.pop(key, val) for key, val in open_kwargs.items()}
            with io.open(path, **open_args) as handle:
                result = handle.read()
                return callback(result, **kwargs) if callable(callback) else result

    else:
        if not callable(callback):
            raise ValueError(f'opener callback {callback!r} not a function')

        # Pass handle to callback
        def method(path, **kwargs):
            open_args = {key: kwargs.pop(key, val) for key, val in open_kwargs.items()}
            with io.open(path, **open_args) as handle:
                return callback(handle, **kwargs)

    return method


@opener
def _markdown(handle, **kwargs):
    from markdown import markdown

    return markdown(handle.read(), **{k: kwargs.pop(k, v) for k, v in _markdown_defaults.items()})


@opener
def _yaml(handle, **kwargs):
    import yaml

    kwargs.setdefault('Loader', yaml.SafeLoader)
    # B506:yaml_load we load safely using SafeLoader
    return yaml.load(handle.read(), **kwargs)  # nosec B506


def _template(path, **kwargs):
    root, name = os.path.split(path)
    return tornado.template.Loader(root, **kwargs).load(name)


# gramex.cache.open() stores its cache here as {(path, callback): {data: ..., stat: ...}}
_OPEN_CACHE = {}
# If _OPEN_CACHE is a Redis/Disk/... cache that can't store the object, use fallback memory cache
_FALLBACK_MEMORY_CACHE = {}
open_callback = {
    'bin': opener(None, read=True, mode='rb', encoding=None, errors=None),
    'txt': opener(None, read=True),
    'text': opener(None, read=True),
    'csv': pd.read_csv,
    'excel': read_excel,
    'xls': pd.read_excel,
    'xlsx': read_excel,
    'hdf': pd.read_hdf,
    'h5': pd.read_hdf,
    'html': pd.read_html,
    'json': opener(json.load),
    'jsondata': pd.read_json,
    'sas': pd.read_sas,
    'stata': pd.read_stata,
    'table': pd.read_table,
    'parquet': pd.read_parquet,
    'feather': pd.read_feather,
    'md': _markdown,
    'markdown': _markdown,
    'tmpl': _template,
    'template': _template,
    'config': PathConfig,
    'yml': _yaml,
    'yaml': _yaml,
}


def _relpath(path):
    # Returns absolute path relative to the caller's caller's frame. Used in open() / save()
    # Get all parent frames, with 2 lines of context in each
    stack = inspect.getouterframes(inspect.currentframe(), context=2)
    # Get the frame that called open() or save(). This is the caller's caller
    frame = stack[2]
    # The folder is relative to that frame's file
    folder = os.path.dirname(os.path.abspath(frame[1]))
    # But if we're calling from a FileHandler template, use the template's path
    if frame[1].endswith('.generated.py'):
        g = frame[0].f_globals
        # Duck-type if it's a FileHandler by checking for .file attribute
        # Don't import gramex.handler.FileHandler -- gramex.cache should have no dependencies
        if 'handler' in g and hasattr(g['handler'], 'file'):
            folder = os.path.dirname(g['handler'].file)
        else:
            app_log.warning(f'gramex.cache.open/save: rel= on unknown template folder for {path}')
    return os.path.join(folder, path)


_SAVE_CALLBACKS = {
    'json': 'to_json',
    'csv': 'to_csv',
    'xlsx': 'to_excel',
    'hdf': 'to_hdf',
    'html': 'to_html',
    'stata': 'to_stata',
    # Other configurations not supported
}


# gramex.cache.query() stores its cache here
_QUERY_CACHE = {}
_STATUS_METHODS = {}


def _wheres(dbkey, tablekey, default_db, names, fn=None):
    '''
    Convert a table name list like ['sales', 'dept.sales']) to a WHERE clause
    like `(table="sales") OR (db="dept" AND table="sales")`.

    TODO: escape the table names to avoid SQL injection attacks
    '''
    where = []
    for name in names:
        db, table = name.rsplit('.', 2) if '.' in name else (default_db, name)
        if not fn:
            where.append("({}='{}' AND {}='{}')".format(dbkey, db, tablekey, table))
        else:
            where.append(
                "({}={}('{}') AND {}={}('{}'))".format(dbkey, fn[0], db, tablekey, fn[1], table)
            )
    return ' OR '.join(where)


def _table_status(engine, tables):
    '''
    Returns the last updated date of a list of tables.
    '''
    # Cache the SQL query or file date check function beforehand.
    # Every time method is called with a URL and table list, run cached query
    dialect = engine.dialect.name
    key = (engine.url, tuple(tables))
    db = engine.url.database
    if _STATUS_METHODS.get(key, None) is None:
        if len(tables) == 0:
            raise ValueError(f'gramex.cache.query table list is empty: {tables!r}')
        for name in tables:
            if not name or not isinstance(name, str):
                raise ValueError(f'gramex.cache.query invalid table list: {tables!r}')
        # bandit security note: We use string substitution for DB and table names.
        # But these are validated via gramex.data._sql_safe, so we're fine.
        if dialect == 'mysql':
            # https://dev.mysql.com/doc/refman/8.0/en/information-schema-tables-table.html
            # Works only on MySQL 5.7 and above
            # B608:hardcoded_sql_expressions only used internally
            w = _wheres('table_schema', 'table_name', db, tables)
            q = 'SELECT update_time FROM information_schema.tables WHERE ' + w  # nosec B608
        elif dialect == 'snowflake':
            # https://docs.snowflake.com/en/sql-reference/info-schema/tables.html
            w = _wheres('table_schema', 'table_name', db, tables)
            q = 'SELECT last_altered FROM information_schema.tables WHERE ' + w  # nosec B608
        elif dialect == 'mssql':
            # https://goo.gl/b4aL9m
            w = _wheres('database_id', 'object_id', db, tables, fn=['DB_ID', 'OBJECT_ID'])
            q = 'SELECT last_user_update FROM sys.dm_db_index_usage_stats WHERE ' + w  # nosec B608
        elif dialect == 'postgresql':
            # https://www.postgresql.org/docs/9.6/static/monitoring-stats.html
            w = _wheres('schemaname', 'relname', 'public', tables)
            q = (
                'SELECT n_tup_ins, n_tup_upd, n_tup_del FROM pg_stat_all_tables '  # nosec B608
                + 'WHERE '
                + w
            )
        elif dialect == 'sqlite':
            if not db:
                raise KeyError(f'gramex.cache.query: does not support memory sqlite "{dialect}"')
            q = db
        else:
            raise KeyError(f'gramex.cache.query: cannot cache dialect "{dialect}" yet')
        if dialect == 'sqlite':
            _STATUS_METHODS[key] = lambda: stat(q)
        else:
            _STATUS_METHODS[key] = lambda: pd.read_sql(q, engine).to_json(orient='records')
    return _STATUS_METHODS[key]()


def _get_current_ioloop():
    '''
    Return the current IOLoop. But if we're not already in an IOLoop, return an
    object that mimics add_callback() by running the method immediately.
    This allows daemon() to be run without Tornado / asyncio.
    '''
    from gramex.services import info

    return (
        info.main_ioloop
        or tornado.ioloop.IOLoop.current()
        or AttrDict(add_callback=lambda fn, *args, **kwargs: fn(*args, **kwargs))
    )


def _create_path(path):
    # Ensure that path directory exists
    path = os.path.abspath(path)
    folder = os.path.dirname(path)
    if not os.path.exists(folder):
        os.makedirs(folder)
    return path


def sizeof(obj):
    if isinstance(obj, dict):
        return sys.getsizeof(obj) + sum(sizeof(k) + sizeof(v) for k, v in obj.items())
    elif isinstance(obj, (set, list)):
        return sys.getsizeof(obj) + sum(sizeof(v) for v in obj)
    return sys.getsizeof(obj)
